import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

def parse_benchmark_results(file_path):
    """Parse the benchmark results from a text file."""
    with open(file_path, 'r') as file:
        content = file.read()
    
    # Dictionary to store all results
    results = {}
    
    # Split content into lines and process
    lines = content.strip().split('\n')
    current_benchmark = None
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Check if this is a benchmark line
        benchmark_match = re.match(r'Benchmark: (.+)', line)
        if benchmark_match:
            current_benchmark = benchmark_match.group(1).strip()
            results[current_benchmark] = {}
            continue
        
        # Check if this is an algorithm result line
        if current_benchmark and ':' in line and 'mean=' in line and 'std=' in line:
            # More flexible pattern to handle various algorithm names
            algorithm_match = re.match(r'(.+?):\s*mean=([^,]+),\s*std=(.+)', line)
            if algorithm_match:
                algorithm = algorithm_match.group(1).strip()
                mean_str = algorithm_match.group(2).strip()
                std_str = algorithm_match.group(3).strip()
                
                try:
                    mean_val = float(mean_str)
                    std_val = float(std_str)
                    results[current_benchmark][algorithm] = {'mean': mean_val, 'std': std_val}
                    print(f"Parsed: {current_benchmark} -> {algorithm}: mean={mean_val}, std={std_val}")
                except ValueError as e:
                    print(f"Warning: Could not parse {algorithm} results for {current_benchmark}: {e}")
                    print(f"  Mean string: '{mean_str}', Std string: '{std_str}'")
    
    return results

def create_excel_file(results, output_file='FCSA_Benchmark_Test3_Dim2_Run1_Results.xlsx'):
    """Create Excel file with the specified format."""
    
    # Get all unique algorithms and benchmarks
    all_algorithms = set()
    benchmarks = list(results.keys())
    
    for benchmark_data in results.values():
        all_algorithms.update(benchmark_data.keys())
    
    all_algorithms = sorted(list(all_algorithms))
    
    print(f"Found {len(all_algorithms)} algorithms across {len(benchmarks)} benchmarks")
    print(f"Algorithms: {all_algorithms}")
    print(f"Benchmarks: {benchmarks}")
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "FCSA Dim 2 Benchmark Results"
    
    # Create headers
    worksheet.cell(row=1, column=1, value="Algorithm").font = Font(bold=True)
    for col, benchmark in enumerate(benchmarks, 2):
        worksheet.cell(row=1, column=col, value=benchmark).font = Font(bold=True)
    
    current_row = 2
    
    # For each algorithm, create two rows (mean and std)
    for algorithm in all_algorithms:
        # Algorithm name in first row, empty in second row for std
        worksheet.cell(row=current_row, column=1, value=algorithm)
        worksheet.cell(row=current_row + 1, column=1, value="")
        
        # Calculate minimum means for each benchmark for bold formatting
        benchmark_mins = {}
        for benchmark in benchmarks:
            if benchmark in results and results[benchmark]:
                means = [data['mean'] for data in results[benchmark].values() if 'mean' in data]
                if means:
                    benchmark_mins[benchmark] = min(means)
        
        # Fill in data for each benchmark
        for col, benchmark in enumerate(benchmarks, 2):
            if benchmark in results and algorithm in results[benchmark]:
                data = results[benchmark][algorithm]
                
                # Mean value (first row)
                mean_cell = worksheet.cell(row=current_row, column=col, value=data['mean'])
                
                # Bold if this is the minimum (best) value for this benchmark
                if benchmark in benchmark_mins:
                    if abs(data['mean'] - benchmark_mins[benchmark]) < 1e-15:
                        mean_cell.font = Font(bold=True)
                
                # Std value (second row)
                worksheet.cell(row=current_row + 1, column=col, value=data['std'])
            else:
                # Algorithm not tested on this benchmark - leave empty
                worksheet.cell(row=current_row, column=col, value="")
                worksheet.cell(row=current_row + 1, column=col, value="")
        
        current_row += 2
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Make algorithm column wider for long names
    worksheet.column_dimensions['A'].width = 35
    
    workbook.save(output_file)
    print(f"Excel file '{output_file}' created successfully!")
    
    return output_file

def main():
    """Main function to run the conversion."""
    # Input file path - change this to your text file path
    input_file = 'test_3_fig\dim_2\Test_2_log.txt'  # Change this to your actual file path

    
    try:
        # Parse the results
        results = parse_benchmark_results(input_file)
        
        # Display summary
        print("Parsed benchmark results:")
        for benchmark, algorithms in results.items():
            print(f"\n{benchmark}: {len(algorithms)} algorithms")
            
            # Find and display the best algorithm for each benchmark
            min_mean = min(data['mean'] for data in algorithms.values())
            best_algorithms = [alg for alg, data in algorithms.items() if abs(data['mean'] - min_mean) < 1e-15]
            print(f"  Best algorithm(s): {', '.join(best_algorithms)} (mean = {min_mean})")
        
        # Create Excel file
        output_file = create_excel_file(results)
        
        print(f"\nConversion completed! Check '{output_file}' for the formatted results.")
        print("Format: Each algorithm has two rows (mean values in first row, std values in second row)")
        print("Best mean values for each benchmark are highlighted in bold.")
        
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        print("Please make sure the file exists and update the 'input_file' variable with the correct path.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()

